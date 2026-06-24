"""
templates_to_excel.py
Converts Essential Model Documentation YAML issue templates → Excel forms.
Output lives in:  <repo_root>/forms/
Run from anywhere:  python scripts/templates_to_excel.py
"""

import os
import re
import glob
import yaml
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ── paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT    = os.path.dirname(SCRIPT_DIR)
TEMPLATE_DIR = os.path.join(REPO_ROOT, ".github", "ISSUE_TEMPLATE")
OUTPUT_DIR   = os.path.join(REPO_ROOT, "forms")

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── colours ────────────────────────────────────────────────────────────────────
C_TITLE_BG    = "1F4E79"   # dark navy  – form title bar
C_TITLE_FG    = "FFFFFF"
C_SECTION_BG  = "D6E4F0"   # light blue – section headers
C_SECTION_FG  = "1F4E79"
C_LABEL_FG    = "1F4E79"   # dark blue  – field labels
C_REQ_FG      = "C00000"   # red        – required asterisk
C_INPUT_BG    = "F2F8FF"   # very light – input cells
C_MULTI_BG    = "FFF8E1"   # amber tint – multi-select note cells
C_BORDER      = "BDD7EE"

# ── helpers ────────────────────────────────────────────────────────────────────
def strip_html_details(text: str) -> str:
    """Remove <details>…</details> blocks (including nested) from a string."""
    while re.search(r'<details[\s\S]*?</details>', text, re.IGNORECASE):
        text = re.sub(r'<details[\s\S]*?</details>', '', text, flags=re.IGNORECASE)
    return text.strip()

def clean_markdown(text: str) -> str:
    """Strip markup; keep readable plain text."""
    if not text:
        return ""
    text = strip_html_details(text)
    text = re.sub(r'#{1,6}\s*(.*)', r'\1', text)          # headings
    text = re.sub(r'\*\*(.*?)\*\*', r'\1', text)           # bold
    text = re.sub(r'\*(.*?)\*',     r'\1', text)           # italic
    text = re.sub(r'`([^`]+)`',     r'\1', text)           # inline code
    text = re.sub(r'!\[.*?\]\(.*?\)', '', text)             # images
    text = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', text)   # links → text
    text = re.sub(r'^\s*[-*]\s+', '• ', text, flags=re.MULTILINE)
    text = re.sub(r'^\s*\|.*\|.*$', '', text, flags=re.MULTILINE)  # tables
    text = re.sub(r'[`>#]', '', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()

def extract_sections(markdown_value: str):
    """
    Split a markdown block into a list of (kind, text) tuples:
      ('divider', '')  for --- lines
      ('h2', text)     for ## Heading
      ('body', text)   for everything else
    Returns a flat list so the caller can render each piece.
    """
    parts = []
    for line in markdown_value.splitlines():
        stripped = line.strip()
        if stripped == '---':
            parts.append(('divider', ''))
        elif re.match(r'^##\s+', stripped):
            parts.append(('h2', re.sub(r'^##\s+', '', stripped)))
        elif re.match(r'^###\s+', stripped):
            parts.append(('h3', re.sub(r'^###\s+', '', stripped)))
    return parts

def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def thin_border(color=C_BORDER):
    side = Side(style="thin", color=color)
    return Border(bottom=side)

# ── per-file builder ────────────────────────────────────────────────────────────
def build_workbook(form: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Form"

    # column widths
    ws.column_dimensions["A"].width = 4    # left margin
    ws.column_dimensions["B"].width = 42   # label
    ws.column_dimensions["C"].width = 4    # gutter
    ws.column_dimensions["D"].width = 38   # input
    ws.column_dimensions["E"].width = 4    # right margin

    row = 1

    # ── title bar ──────────────────────────────────────────────────────────────
    title_text = form.get("name", "Form")
    # strip stage prefix noise (e.g. "Stage 4: ") for cleaner title
    title_text = re.sub(r'^Stage\s+\d+:\s*', '', title_text)

    ws.merge_cells(f"A{row}:E{row}")
    tc = ws.cell(row=row, column=1, value=f"  {title_text}")
    tc.fill   = make_fill(C_TITLE_BG)
    tc.font   = Font(name="Arial", size=15, bold=True, color=C_TITLE_FG)
    tc.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 32
    row += 2

    # ── body items ────────────────────────────────────────────────────────────
    for item in form.get("body", []):
        item_type = item.get("type")
        attrs     = item.get("attributes", {})

        # ── markdown blocks: extract structure ─────────────────────────────
        if item_type == "markdown":
            raw = attrs.get("value", "")
            for kind, text in extract_sections(raw):
                if kind == "divider":
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).border = thin_border("1F4E79")
                    row += 1
                elif kind == "h2":
                    ws.merge_cells(f"A{row}:E{row}")
                    sc = ws.cell(row=row, column=1, value=f"  {text}")
                    sc.fill      = make_fill(C_SECTION_BG)
                    sc.font      = Font(name="Arial", size=11, bold=True, color=C_SECTION_FG)
                    sc.alignment = Alignment(vertical="center", indent=1)
                    ws.row_dimensions[row].height = 22
                    row += 1
                elif kind == "h3":
                    ws.merge_cells(f"B{row}:D{row}")
                    sc = ws.cell(row=row, column=2, value=text)
                    sc.font      = Font(name="Arial", size=10, bold=True, color=C_SECTION_FG)
                    sc.alignment = Alignment(vertical="center")
                    row += 1
            row += 1
            continue

        if "id" not in item:
            continue

        label    = clean_markdown(attrs.get("label", item["id"]))
        options  = attrs.get("options", [])
        required = item.get("validations", {}).get("required", False)
        multiple = attrs.get("multiple", False)

        # ── label cell ────────────────────────────────────────────────────
        lc = ws.cell(row=row, column=2, value=label)
        lc.font      = Font(name="Arial", size=10, bold=True, color=C_LABEL_FG)
        lc.alignment = Alignment(vertical="top", wrap_text=True)

        if required:
            rc = ws.cell(row=row, column=1, value="*")
            rc.font      = Font(name="Arial", bold=True, color=C_REQ_FG)
            rc.alignment = Alignment(horizontal="center", vertical="top")

        # ── input cell ────────────────────────────────────────────────────
        ic = ws.cell(row=row, column=4)
        ic.alignment = Alignment(wrap_text=True, vertical="top")

        if multiple:
            ic.fill = make_fill(C_MULTI_BG)
            ic.font = Font(name="Arial", size=9, color="7B6000", italic=True)
            ic.value = "(select multiple — enter comma-separated values)"
        elif options:
            ic.fill = make_fill(C_INPUT_BG)
            formula_str = ",".join(str(o) for o in options)
            if len(formula_str) <= 255:
                dv = DataValidation(
                    type="list",
                    formula1=f'"{formula_str}"',
                    allow_blank=True,
                    showErrorMessage=True,
                    error="Please select a value from the list.",
                    errorTitle="Invalid entry"
                )
                ws.add_data_validation(dv)
                dv.add(ic)
        elif item_type == "textarea":
            ic.fill = make_fill(C_INPUT_BG)
            ws.row_dimensions[row].height = 60
        else:
            ic.fill = make_fill(C_INPUT_BG)

        # subtle row divider
        ws.cell(row=row + 1, column=2).border = thin_border()
        row += 2

    # ── freeze title ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    return wb

# ── main ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    yaml_files = sorted(glob.glob(os.path.join(TEMPLATE_DIR, "*.yml")))
    processed  = 0

    for path in yaml_files:
        fname = os.path.basename(path)
        if fname == "config.yml":
            continue

        with open(path, "r", encoding="utf-8") as f:
            form = yaml.safe_load(f)

        wb = build_workbook(form)

        # output filename = template stem, no suffix clutter
        stem   = os.path.splitext(fname)[0]
        output = os.path.join(OUTPUT_DIR, f"{stem}.xlsx")
        wb.save(output)
        print(f"  ✓  {fname:45s} → forms/{stem}.xlsx")
        processed += 1

    print(f"\nDone — {processed} form(s) written to {OUTPUT_DIR}/")
